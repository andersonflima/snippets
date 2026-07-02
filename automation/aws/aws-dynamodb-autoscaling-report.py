#!/usr/bin/env python3
"""Relatório Excel de DynamoDB com Application Auto Scaling ATIVO por conta.

Para cada conta listada no CSV, o script assume uma role (STS AssumeRole) e, na
região sa-east-1, lê os Application Auto Scaling *scalable targets* do namespace
`dynamodb`. Cada target existente representa uma dimensão (leitura ou escrita) de
uma tabela ou de um GSI que TEM auto scaling ativo — só isso entra no relatório.

As credenciais de origem (a conta a partir da qual se assume a role) vêm
automaticamente do ambiente AWS padrão: variáveis AWS_* ou o profile em uso
(a mesma cadeia de resolução do boto3/CLI). Nada de credencial hardcoded.

Uso:
    pip install boto3 openpyxl
    export AWS_PROFILE=minha-conta-bastion   # ou AWS_ACCESS_KEY_ID/SECRET/...
    python aws-dynamodb-autoscaling-report.py \
        --accounts-csv contas.csv \
        --assume-role OrgReadOnly

O CSV aceita cabeçalho com as colunas `account_id` e (opcional) `account_name`;
sem cabeçalho reconhecido, usa a 1ª coluna como id e a 2ª como nome.

Saída: dynamodb-autoscaling-<timestamp>.xlsx com uma aba `autoscaling`
(recursos com auto scaling ativo) e uma aba `falhas` (contas não processadas).
"""

from __future__ import annotations

import argparse
import csv
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional

import boto3
from botocore.exceptions import BotoCoreError, ClientError

REGION = "sa-east-1"

DIMENSION_LABELS = {
    "dynamodb:table:ReadCapacityUnits": "leitura",
    "dynamodb:table:WriteCapacityUnits": "escrita",
    "dynamodb:index:ReadCapacityUnits": "leitura",
    "dynamodb:index:WriteCapacityUnits": "escrita",
}


# --------------------------------------------------------------------------- #
# CLI                                                                         #
# --------------------------------------------------------------------------- #
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Gera relatório Excel das tabelas DynamoDB com Application Auto "
            f"Scaling ATIVO em {REGION}, varrendo várias contas via assume-role."
        )
    )
    parser.add_argument(
        "--accounts-csv",
        required=True,
        type=Path,
        help="CSV com as contas (coluna account_id e, opcional, account_name).",
    )
    parser.add_argument(
        "--assume-role",
        required=True,
        help="Nome da role a assumir em cada conta (ex: OrgReadOnly).",
    )
    parser.add_argument(
        "--external-id",
        help="ExternalId opcional exigido pela trust policy da role.",
    )
    parser.add_argument(
        "--role-session-name",
        default="dynamodb-autoscaling-report",
        help="SessionName do AssumeRole (padrão: dynamodb-autoscaling-report).",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=8,
        help="Contas processadas em paralelo (padrão: 8).",
    )
    parser.add_argument(
        "--report",
        help="Caminho do .xlsx de saída (padrão: dynamodb-autoscaling-<timestamp>.xlsx).",
    )
    return parser.parse_args()


# --------------------------------------------------------------------------- #
# Carga de contas                                                             #
# --------------------------------------------------------------------------- #
def _normalize_header(header: str) -> str:
    return "".join(char for char in header.strip().lower() if char.isalnum())


_ID_HEADERS = {"accountid", "account", "conta", "contaid", "id", "numerodaconta"}
_NAME_HEADERS = {"accountname", "name", "nome", "nomedaconta", "conta", "alias"}


def _account(account_id: str, account_name: str = "") -> dict:
    return {"account_id": account_id.strip(), "account_name": account_name.strip()}


def load_accounts(accounts_csv: Path) -> List[dict]:
    with accounts_csv.open(encoding="utf-8-sig", newline="") as handler:
        rows = [row for row in csv.reader(handler) if any(cell.strip() for cell in row)]
    if not rows:
        raise ValueError("CSV de contas vazio.")

    headers = [_normalize_header(cell) for cell in rows[0]]
    id_index, name_index, data_rows = 0, 1, rows
    if any(header in _ID_HEADERS for header in headers):
        id_index = next(i for i, h in enumerate(headers) if h in _ID_HEADERS)
        name_index = next((i for i, h in enumerate(headers) if h in _NAME_HEADERS and i != id_index), None)
        data_rows = rows[1:]

    accounts, seen = [], set()
    for row in data_rows:
        if len(row) <= id_index or not row[id_index].strip():
            continue
        account_id = row[id_index].strip()
        if account_id in seen:
            continue
        seen.add(account_id)
        name = row[name_index].strip() if name_index is not None and len(row) > name_index else ""
        accounts.append(_account(account_id, name))

    if not accounts:
        raise ValueError("Nenhuma conta válida encontrada no CSV.")
    return accounts


# --------------------------------------------------------------------------- #
# AWS                                                                         #
# --------------------------------------------------------------------------- #
def assume_role_session(
    source: boto3.Session,
    account_id: str,
    role_name: str,
    role_session_name: str,
    external_id: Optional[str],
) -> boto3.Session:
    sts = source.client("sts", region_name=REGION)
    params = {
        "RoleArn": f"arn:aws:iam::{account_id}:role/{role_name}",
        "RoleSessionName": role_session_name,
    }
    if external_id:
        params["ExternalId"] = external_id

    creds = sts.assume_role(**params)["Credentials"]
    return boto3.Session(
        aws_access_key_id=creds["AccessKeyId"],
        aws_secret_access_key=creds["SecretAccessKey"],
        aws_session_token=creds["SessionToken"],
        region_name=REGION,
    )


def _parse_resource_id(resource_id: str) -> tuple[str, str, str]:
    """Ex: 'table/Foo' -> (Foo, tabela, ''); 'table/Foo/index/Bar' -> (Foo, GSI, Bar)."""
    parts = resource_id.split("/")
    table = parts[1] if len(parts) > 1 else resource_id
    if len(parts) >= 4 and parts[2] == "index":
        return table, "GSI", parts[3]
    return table, "tabela", ""


def scan_autoscaling(session: boto3.Session) -> List[dict]:
    """Só retorna dimensões com auto scaling ativo (targets existentes)."""
    client = session.client("application-autoscaling", region_name=REGION)
    paginator = client.get_paginator("describe_scalable_targets")

    rows: List[dict] = []
    for page in paginator.paginate(ServiceNamespace="dynamodb"):
        for target in page.get("ScalableTargets", []):
            table, scope, index_name = _parse_resource_id(target["ResourceId"])
            rows.append(
                {
                    "table": table,
                    "scope": scope,
                    "index_name": index_name,
                    "dimension": DIMENSION_LABELS.get(
                        target["ScalableDimension"], target["ScalableDimension"]
                    ),
                    "min_capacity": target.get("MinCapacity"),
                    "max_capacity": target.get("MaxCapacity"),
                    "resource_id": target["ResourceId"],
                }
            )
    return rows


def log(message: str) -> None:
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    print(f"[{stamp}] {message}", file=sys.stderr, flush=True)


def check_account(
    index: int,
    account: dict,
    source: boto3.Session,
    role_name: str,
    role_session_name: str,
    external_id: Optional[str],
) -> tuple[int, dict]:
    account_id = account["account_id"]
    result = {**account, "ok": False, "error": "", "error_code": "", "resources": []}
    try:
        log(f"Conta {index + 1} ({account_id}): assumindo role")
        session = assume_role_session(source, account_id, role_name, role_session_name, external_id)
        resources = scan_autoscaling(session)
        result["resources"] = resources
        result["ok"] = True
        log(f"Conta {account_id}: recursos_com_autoscaling={len(resources)}")
    except (ClientError, BotoCoreError, ValueError) as error:
        if isinstance(error, ClientError):
            result["error_code"] = error.response.get("Error", {}).get("Code", "")
        result["error"] = str(error)
        log(f"Conta {account_id}: ERRO {result['error_code']} {result['error']}")
    return index, result


# --------------------------------------------------------------------------- #
# Relatório                                                                   #
# --------------------------------------------------------------------------- #
AUTOSCALING_HEADERS = [
    "account_id",
    "account_name",
    "regiao",
    "recurso",
    "tipo",
    "indice",
    "dimensao",
    "min_capacity",
    "max_capacity",
    "resource_id",
]

FAILED_HEADERS = ["account_id", "account_name", "error_code", "error"]


def _autoscaling_rows(results: List[dict]) -> List[list]:
    rows: List[list] = []
    for item in results:
        for resource in item["resources"]:
            rows.append(
                [
                    item["account_id"],
                    item["account_name"],
                    REGION,
                    resource["table"],
                    resource["scope"],
                    resource["index_name"],
                    resource["dimension"],
                    resource["min_capacity"],
                    resource["max_capacity"],
                    resource["resource_id"],
                ]
            )
    return rows


def _account_label(item: dict) -> str:
    name = item["account_name"]
    return f"{item['account_id']} ({name})" if name else item["account_id"]


def print_console_summary(results: List[dict], report_path: str) -> None:
    """Resumo claro no stdout, destacando as contas que falharam."""
    failed = [item for item in results if not item["ok"]]
    ok_count = len(results) - len(failed)
    total_resources = sum(len(item["resources"]) for item in results)

    print(f"\nRelatório: {report_path}")
    print(
        f"Contas OK: {ok_count}/{len(results)} | "
        f"recursos com auto scaling: {total_resources} | "
        f"contas com erro: {len(failed)}"
    )
    if failed:
        print(f"\n{len(failed)} conta(s) NÃO processada(s) (também na aba 'falhas'):")
        for item in failed:
            reason = item["error_code"] or item["error"] or "erro desconhecido"
            print(f"  - {_account_label(item)}: {reason}")


def write_report(report_path: str, results: List[dict]) -> str:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)

    sheet = workbook.create_sheet("autoscaling")
    sheet.append(AUTOSCALING_HEADERS)
    for row in _autoscaling_rows(results):
        sheet.append(row)

    failed_sheet = workbook.create_sheet("falhas")
    failed_sheet.append(FAILED_HEADERS)
    for item in results:
        if not item["ok"]:
            failed_sheet.append(
                [item["account_id"], item["account_name"], item["error_code"], item["error"]]
            )

    path = Path(report_path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    workbook.save(path)
    return str(path)


# --------------------------------------------------------------------------- #
# Orquestração                                                                #
# --------------------------------------------------------------------------- #
def run(args: argparse.Namespace) -> int:
    if args.workers <= 0:
        raise ValueError("--workers precisa ser maior que 0.")

    accounts = load_accounts(args.accounts_csv)
    source = boto3.Session()  # credenciais AWS resolvidas automaticamente do ambiente
    if source.get_credentials() is None:
        raise ValueError(
            "Credenciais AWS não encontradas no ambiente. Configure AWS_PROFILE ou "
            "AWS_ACCESS_KEY_ID/AWS_SECRET_ACCESS_KEY antes de rodar."
        )

    workers = min(args.workers, len(accounts))
    log(f"Iniciando: contas={len(accounts)} workers={workers} role={args.assume_role} regiao={REGION}")

    results: List[dict] = [None] * len(accounts)  # type: ignore[list-item]
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [
            executor.submit(
                check_account,
                index=index,
                account=account,
                source=source,
                role_name=args.assume_role,
                role_session_name=args.role_session_name,
                external_id=args.external_id,
            )
            for index, account in enumerate(accounts)
        ]
        for future in as_completed(futures):
            index, result = future.result()
            results[index] = result

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    report_path = write_report(args.report or f"dynamodb-autoscaling-{timestamp}.xlsx", results)

    total_resources = sum(len(item["resources"]) for item in results)
    failed = [item for item in results if not item["ok"]]
    log(
        f"Concluído: relatorio={report_path} recursos_com_autoscaling={total_resources} "
        f"contas_ok={len(results) - len(failed)} contas_falha={len(failed)}"
    )
    print_console_summary(results, report_path)
    return 2 if failed else 0


def main() -> int:
    return run(parse_args())


if __name__ == "__main__":
    raise SystemExit(main())
