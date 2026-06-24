#!/usr/bin/env python3
"""Audita DynamoDB com Application Auto Scaling em várias contas (assume-role).

Para cada conta-alvo (assumindo uma role comum a partir das credenciais da conta
bastion no ambiente) e em cada região, o script:

1. Lista as tabelas DynamoDB e descreve billing mode + GSIs.
2. Lê os Application Auto Scaling *scalable targets* (ServiceNamespace=dynamodb)
   e mapeia, por tabela, quais dimensões têm auto scaling:
   - tabela:  ReadCapacityUnits / WriteCapacityUnits
   - cada GSI: ReadCapacityUnits / WriteCapacityUnits
3. Deriva quais atributos precisam estar no `lifecycle.ignore_changes` do
   Terraform (read_capacity / write_capacity da tabela e de cada GSI) — porque
   o auto scaling altera essas capacities fora do Terraform.
4. Classifica os CENÁRIOS de cada tabela:
   - `on_demand`                  PAY_PER_REQUEST (capacity não se aplica)
   - `provisioned_no_autoscaling` provisioned sem nenhum target
   - `table_partial_autoscaling`  só read OU só write da tabela tem auto scaling
   - `gsi_partial:<idx>`          GSI com só read OU só write
   - `fully_autoscaled`           todas as dimensões existentes têm auto scaling

Observação: `ignore_changes` é um conceito do Terraform; a API da AWS não o
expõe. O script reporta a REALIDADE na AWS (dimensões com auto scaling) para que
você verifique se o `ignore_changes` do seu Terraform cobre todos esses casos.

Saída: relatório Excel com as abas:
  - `autoscaling` : só tabelas que têm auto scaling + ignore_changes requerido.
  - `tables`      : todas as tabelas avaliadas (todos os cenários).
  - `failed`      : contas que não puderam ser processadas.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from os import getenv
from pathlib import Path
from typing import List, Optional, Set

import boto3
from botocore.exceptions import BotoCoreError, ClientError

TABLE_READ = "dynamodb:table:ReadCapacityUnits"
TABLE_WRITE = "dynamodb:table:WriteCapacityUnits"
INDEX_READ = "dynamodb:index:ReadCapacityUnits"
INDEX_WRITE = "dynamodb:index:WriteCapacityUnits"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Procura tabelas DynamoDB com Application Auto Scaling em múltiplas "
            "contas (assume-role) e verifica, por cenário, quais capacities "
            "precisam estar no lifecycle.ignore_changes do Terraform."
        )
    )
    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument(
        "--accounts",
        help="Lista de contas separadas por vírgula (ex: 111111111111,222222222222).",
    )
    target.add_argument(
        "--accounts-file",
        type=Path,
        help="Arquivo com um account id por linha.",
    )
    target.add_argument(
        "--accounts-csv",
        type=Path,
        help="Arquivo CSV com a coluna `account_id` (ou `account`, `accountId`, primeira coluna).",
    )
    parser.add_argument(
        "--assume-role",
        required=True,
        help="Nome da role a assumir em cada conta-alvo (ex: OrgReadOnly).",
    )
    parser.add_argument(
        "--region",
        default="us-east-1",
        help="Uma ou mais regiões separadas por vírgula (padrão: us-east-1).",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=8,
        help="Quantidade máxima de contas processadas em paralelo.",
    )
    parser.add_argument(
        "--report",
        dest="report",
        help=(
            "Arquivo Excel de saída com abas autoscaling/tables/failed "
            "(padrão: dynamodb-autoscaling-report-<timestamp>.xlsx)."
        ),
    )
    parser.add_argument(
        "--external-id",
        help="ExternalId opcional para o AssumeRole.",
    )
    parser.add_argument(
        "--role-session-name",
        default="dynamodb-autoscaling-check",
        help="SessionName do AssumeRole (padrão: dynamodb-autoscaling-check).",
    )
    parser.add_argument(
        "--fail-on-partial",
        action="store_true",
        help="Sai com código 3 se alguma tabela tiver auto scaling parcial/assimétrico.",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Imprime também a saída detalhada em JSON.",
    )
    return parser.parse_args()


# --------------------------------------------------------------------------- #
# Account loading                                                             #
# --------------------------------------------------------------------------- #
def _account_record(account_id: str) -> dict:
    return {"account": account_id}


def _normalize_header(header: str) -> str:
    return "".join(char for char in header.strip().lower() if char.isalnum())


def _account_id_headers() -> Set[str]:
    return {"account", "accountid", "accountnumber", "id", "conta", "contaid", "numerodaconta"}


def _load_accounts_from_csv(accounts_csv: Path) -> List[dict]:
    with accounts_csv.open(encoding="utf-8-sig", newline="") as handler:
        rows = list(csv.reader(handler))
    if not rows or not rows[0]:
        return []

    headers = [_normalize_header(cell) for cell in rows[0]]
    if any(header in _account_id_headers() for header in headers):
        account_index = next(
            index for index, header in enumerate(headers) if header in _account_id_headers()
        )
        return [
            _account_record(row[account_index].strip())
            for row in rows[1:]
            if len(row) > account_index and row[account_index].strip()
        ]

    return [_account_record(row[0].strip()) for row in rows if row and row[0].strip()]


def _dedupe_accounts(values: List[dict]) -> List[dict]:
    seen: Set[str] = set()
    deduped: List[dict] = []
    for value in values:
        account_id = value["account"]
        if account_id and account_id not in seen:
            seen.add(account_id)
            deduped.append(value)
    return deduped


def _account_load_result(values: List[dict]) -> dict:
    valid_accounts = [value for value in values if value["account"]]
    unique_accounts = _dedupe_accounts(valid_accounts)
    return {
        "accounts": unique_accounts,
        "raw_count": len(values),
        "valid_count": len(valid_accounts),
        "empty_count": len(values) - len(valid_accounts),
        "duplicate_count": len(valid_accounts) - len(unique_accounts),
    }


def load_accounts(
    accounts_csv: Optional[str],
    accounts_file: Optional[Path],
    accounts_csv_file: Optional[Path],
) -> dict:
    if accounts_csv:
        values = [_account_record(acc.strip()) for acc in accounts_csv.split(",")]
    elif accounts_file:
        values = [_account_record(line.strip()) for line in accounts_file.read_text().splitlines()]
    elif accounts_csv_file:
        values = _load_accounts_from_csv(accounts_csv_file)
    else:
        raise ValueError("Informe --accounts, --accounts-file ou --accounts-csv.")

    result = _account_load_result(values)
    if not result["accounts"]:
        raise ValueError("Nenhuma conta válida encontrada.")
    return result


def parse_regions(raw: str) -> List[str]:
    regions = [item.strip() for item in raw.split(",") if item.strip()]
    if not regions:
        raise ValueError("Informe ao menos uma região em --region.")
    deduped: List[str] = []
    for region in regions:
        if region not in deduped:
            deduped.append(region)
    return deduped


# --------------------------------------------------------------------------- #
# Source (bastion) credentials                                                #
# --------------------------------------------------------------------------- #
def _read_json_path_or_text(raw: str) -> dict:
    raw = raw.strip()
    if not raw:
        return {}

    path = Path(raw)
    if path.exists():
        raw = path.read_text(encoding="utf-8")

    data = json.loads(raw)
    if not isinstance(data, dict):
        raise ValueError("AWS_SECRETS deve ser um JSON válido de mapa.")
    return data


def _resolve_source_credentials() -> dict:
    aws_secrets = getenv("AWS_SECRETS")
    if aws_secrets:
        secrets = _read_json_path_or_text(aws_secrets)
        return {
            "aws_access_key_id": (
                secrets.get("aws_access_key_id")
                or secrets.get("AccessKeyId")
                or secrets.get("accessKeyId")
                or getenv("AWS_ACCESS_KEY_ID")
            ),
            "aws_secret_access_key": (
                secrets.get("aws_secret_access_key")
                or secrets.get("SecretAccessKey")
                or secrets.get("secretAccessKey")
                or getenv("AWS_SECRET_ACCESS_KEY")
            ),
            "aws_session_token": (
                secrets.get("aws_session_token")
                or secrets.get("SessionToken")
                or secrets.get("sessionToken")
                or getenv("AWS_SESSION_TOKEN")
            ),
            "region": (
                secrets.get("region")
                or secrets.get("aws_region")
                or secrets.get("AWS_REGION")
                or getenv("AWS_REGION")
            ),
        }

    return {
        "aws_access_key_id": getenv("AWS_ACCESS_KEY_ID"),
        "aws_secret_access_key": getenv("AWS_SECRET_ACCESS_KEY"),
        "aws_session_token": getenv("AWS_SESSION_TOKEN"),
        "region": getenv("AWS_REGION") or getenv("AWS_DEFAULT_REGION"),
    }


def build_source_session(region_name: str) -> boto3.Session:
    credentials = _resolve_source_credentials()
    access_key_id = credentials.get("aws_access_key_id")
    secret_access_key = credentials.get("aws_secret_access_key")

    if not access_key_id or not secret_access_key:
        raise ValueError(
            "Credenciais da conta bastion não encontradas. Defina AWS_SECRETS ou "
            "AWS_ACCESS_KEY_ID + AWS_SECRET_ACCESS_KEY no ambiente."
        )

    return boto3.Session(
        aws_access_key_id=access_key_id,
        aws_secret_access_key=secret_access_key,
        aws_session_token=credentials.get("aws_session_token"),
        region_name=credentials.get("region") or region_name,
    )


# --------------------------------------------------------------------------- #
# AWS work per account                                                        #
# --------------------------------------------------------------------------- #
def assume_role_for_account(
    source_session: boto3.Session,
    account_id: str,
    role_name: str,
    role_session_name: str,
    external_id: Optional[str],
    region: str,
) -> boto3.Session:
    sts = source_session.client("sts", region_name=region)
    params = {
        "RoleArn": f"arn:aws:iam::{account_id}:role/{role_name}",
        "RoleSessionName": role_session_name,
    }
    if external_id:
        params["ExternalId"] = external_id

    creds = sts.assume_role(**params)["Credentials"]
    return boto3.Session(
        aws_access_key_id=creds["AccessKeyId"],
        aws_secret_access_key=creds["SecretAccessKey"],
        aws_session_token=creds["SessionToken"],
        region_name=region,
    )


def _scalable_targets(assumed_session: boto3.Session, region: str) -> dict:
    """Mapa resourceId -> {scalableDimension} para o namespace dynamodb."""
    client = assumed_session.client("application-autoscaling", region_name=region)
    targets: dict[str, Set[str]] = {}
    paginator = client.get_paginator("describe_scalable_targets")
    for page in paginator.paginate(ServiceNamespace="dynamodb"):
        for target in page.get("ScalableTargets", []):
            targets.setdefault(target["ResourceId"], set()).add(target["ScalableDimension"])
    return targets


def _evaluate_table(name: str, desc: dict, targets: dict, region: str) -> dict:
    billing = (desc.get("BillingModeSummary") or {}).get("BillingMode") or "PROVISIONED"
    gsis = [gsi["IndexName"] for gsi in desc.get("GlobalSecondaryIndexes", [])]

    table_dims = targets.get(f"table/{name}", set())
    table_read = TABLE_READ in table_dims
    table_write = TABLE_WRITE in table_dims

    gsi_cov = {}
    for index_name in gsis:
        dims = targets.get(f"table/{name}/index/{index_name}", set())
        gsi_cov[index_name] = {"read": INDEX_READ in dims, "write": INDEX_WRITE in dims}

    autoscaled: List[str] = []
    required: List[str] = []
    if table_read:
        autoscaled.append("table:read")
        required.append("read_capacity")
    if table_write:
        autoscaled.append("table:write")
        required.append("write_capacity")
    for index_name, cov in gsi_cov.items():
        if cov["read"]:
            autoscaled.append(f"index:{index_name}:read")
            required.append(f'global_secondary_index["{index_name}"].read_capacity')
        if cov["write"]:
            autoscaled.append(f"index:{index_name}:write")
            required.append(f'global_secondary_index["{index_name}"].write_capacity')

    has_autoscaling = bool(autoscaled)

    scenarios: List[str] = []
    if billing == "PAY_PER_REQUEST":
        scenarios.append("on_demand")
    else:
        if not has_autoscaling:
            scenarios.append("provisioned_no_autoscaling")
        if table_read != table_write:
            scenarios.append("table_partial_autoscaling")
        for index_name, cov in gsi_cov.items():
            if cov["read"] != cov["write"]:
                scenarios.append(f"gsi_partial:{index_name}")
        if (
            has_autoscaling
            and table_read
            and table_write
            and all(cov["read"] and cov["write"] for cov in gsi_cov.values())
        ):
            scenarios.append("fully_autoscaled")

    partial = any(s == "table_partial_autoscaling" or s.startswith("gsi_partial:") for s in scenarios)

    return {
        "region": region,
        "table": name,
        "billing_mode": billing,
        "gsi_count": len(gsis),
        "has_autoscaling": has_autoscaling,
        "autoscaled_dimensions": autoscaled,
        "needs_ignore_changes": has_autoscaling,
        "required_ignore_changes": required,
        "scenarios": scenarios,
        "partial": partial,
    }


def scan_region(assumed_session: boto3.Session, region: str) -> List[dict]:
    dynamodb = assumed_session.client("dynamodb", region_name=region)
    targets = _scalable_targets(assumed_session, region)

    rows: List[dict] = []
    paginator = dynamodb.get_paginator("list_tables")
    for page in paginator.paginate():
        for name in page.get("TableNames", []):
            desc = dynamodb.describe_table(TableName=name)["Table"]
            rows.append(_evaluate_table(name, desc, targets, region))
    return rows


def _extract_error_code(error: Exception) -> str:
    if not isinstance(error, ClientError):
        return ""
    return (error.response or {}).get("Error", {}).get("Code", "")


def log_step(message: str) -> None:
    print(f"[{datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')}] {message}", file=sys.stderr, flush=True)


def check_account(
    index: int,
    account: dict,
    args: argparse.Namespace,
    source_session: boto3.Session,
    regions: List[str],
) -> tuple[int, dict]:
    account_id = account["account"]
    result = {
        "account": account_id,
        "regions": regions,
        "ok": False,
        "error": None,
        "error_code": None,
        "tables": [],
    }
    try:
        log_step(f"Conta {index + 1}: iniciando {account_id}")
        assumed_session = assume_role_for_account(
            source_session=source_session,
            account_id=account_id,
            role_name=args.assume_role,
            role_session_name=args.role_session_name,
            external_id=args.external_id,
            region=regions[0],
        )
        log_step(f"Conta {account_id}: assumeRole concluido")

        tables: List[dict] = []
        for region in regions:
            region_rows = scan_region(assumed_session, region)
            for row in region_rows:
                row["account"] = account_id
            tables.extend(region_rows)
            log_step(
                f"Conta {account_id} [{region}]: tabelas={len(region_rows)} "
                f"com_autoscaling={sum(1 for r in region_rows if r['has_autoscaling'])}"
            )
        result["tables"] = tables
        result["ok"] = True
    except (ClientError, BotoCoreError, ValueError) as error:
        result["error_code"] = _extract_error_code(error) or None
        result["error"] = str(error)
        result["ok"] = False
        log_step(f"Conta {account_id}: erro: {result['error']}")
    return index, result


# --------------------------------------------------------------------------- #
# Reporting                                                                   #
# --------------------------------------------------------------------------- #
def _yes_no(value: bool) -> str:
    return "sim" if value else "nao"


def table_headers() -> List[str]:
    return [
        "account",
        "region",
        "table",
        "billing_mode",
        "has_autoscaling",
        "autoscaled_dimensions",
        "needs_ignore_changes",
        "required_ignore_changes",
        "scenarios",
        "partial",
    ]


def table_row(item: dict) -> List[str]:
    return [
        item["account"],
        item["region"],
        item["table"],
        item["billing_mode"],
        _yes_no(item["has_autoscaling"]),
        ", ".join(item["autoscaled_dimensions"]),
        _yes_no(item["needs_ignore_changes"]),
        ", ".join(item["required_ignore_changes"]),
        ", ".join(item["scenarios"]),
        _yes_no(item["partial"]),
    ]


def failed_headers() -> List[str]:
    return ["account", "error_code", "error"]


def failed_row(item: dict) -> List[str]:
    return [item["account"], item["error_code"] or "", item["error"] or ""]


def all_tables(results: List[dict]) -> List[dict]:
    rows: List[dict] = []
    for item in results:
        if item["ok"]:
            rows.extend(item["tables"])
    return rows


def write_xlsx_report(report_path: str, results: List[dict]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as error:
        raise ValueError("Para gerar relatório Excel, instale openpyxl.") from error

    tables = all_tables(results)
    autoscaling = [row for row in tables if row["has_autoscaling"]]
    failed = [item for item in results if not item["ok"]]

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, rows in (("autoscaling", autoscaling), ("tables", tables)):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(table_headers())
        for row in rows:
            sheet.append(table_row(row))

    failed_sheet = workbook.create_sheet("failed")
    failed_sheet.append(failed_headers())
    for item in failed:
        failed_sheet.append(failed_row(item))

    workbook.save(report_path)


def normalize_report_path(report_path: str) -> str:
    path = Path(report_path)
    if path.suffix.lower() == ".xlsx":
        return str(path)
    return str(path.with_suffix(".xlsx"))


def write_reports(report_path: str, results: List[dict]) -> str:
    report_path = normalize_report_path(report_path)
    write_xlsx_report(report_path, results)

    tables = all_tables(results)
    autoscaling = sum(1 for row in tables if row["has_autoscaling"])
    failed = sum(1 for item in results if not item["ok"])
    log_step(
        f"Relatorio Excel gerado: {report_path} sheets=autoscaling,tables,failed "
        f"tabelas={len(tables)} com_autoscaling={autoscaling} contas_falhas={failed}"
    )
    return report_path


# --------------------------------------------------------------------------- #
# Orchestration                                                               #
# --------------------------------------------------------------------------- #
def run_check(args: argparse.Namespace) -> int:
    if args.workers <= 0:
        raise ValueError("workers precisa ser maior que 0.")

    regions = parse_regions(args.region)
    account_load = load_accounts(args.accounts, args.accounts_file, args.accounts_csv)
    accounts = account_load["accounts"]
    source_session = build_source_session(regions[0])
    workers = min(args.workers, len(accounts))

    log_step(
        "Carga de contas: "
        f"entradas_lidas={account_load['raw_count']} "
        f"validas={account_load['valid_count']} "
        f"vazias_ignoradas={account_load['empty_count']} "
        f"duplicadas_removidas={account_load['duplicate_count']} "
        f"unicas_processadas={len(accounts)}"
    )
    log_step(
        f"Iniciando auditoria DynamoDB auto scaling: total_contas={len(accounts)} "
        f"workers={workers} assume_role={args.assume_role} regioes={','.join(regions)}"
    )

    results: list[dict] = [None for _ in accounts]  # type: ignore[list-item]
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [
            executor.submit(
                check_account,
                index=index,
                account=account,
                args=args,
                source_session=source_session,
                regions=regions,
            )
            for index, account in enumerate(accounts)
        ]
        for future in as_completed(futures):
            index, result = future.result()
            results[index] = result

    report_path = args.report or f"dynamodb-autoscaling-report-{datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}.xlsx"
    write_reports(report_path, results)

    tables = all_tables(results)
    autoscaling_rows = [row for row in tables if row["has_autoscaling"]]
    partial_rows = [row for row in tables if row["partial"]]
    errors = [item for item in results if not item["ok"]]

    if args.json:
        print(json.dumps(results, indent=2, ensure_ascii=False))
    else:
        for item in results:
            if not item["ok"]:
                print(f'{item["account"]}: ERRO - {item["error"] or "erro desconhecido"}')
                continue
            with_as = sum(1 for row in item["tables"] if row["has_autoscaling"])
            partial = sum(1 for row in item["tables"] if row["partial"])
            print(
                f'{item["account"]}: OK - tabelas={len(item["tables"])} '
                f'com_auto_scaling={with_as} parciais={partial}'
            )
            for row in item["tables"]:
                if row["partial"]:
                    print(
                        f'    [PARCIAL] {row["region"]}/{row["table"]}: '
                        f'cenarios={", ".join(row["scenarios"])} '
                        f'ignore_changes={", ".join(row["required_ignore_changes"]) or "-"}'
                    )

    log_step(
        f"Concluido: tabelas={len(tables)} com_auto_scaling={len(autoscaling_rows)} "
        f"parciais={len(partial_rows)} contas_com_erro={len(errors)}"
    )

    if errors:
        return 2
    if args.fail_on_partial and partial_rows:
        return 3
    return 0


def main() -> int:
    args = parse_args()
    return run_check(args)


if __name__ == "__main__":
    raise SystemExit(main())
