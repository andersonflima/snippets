#!/usr/bin/env python3
"""Remove snapshots MANUAIS de RDS de um recurso, em várias contas (assume-role).

Resource-driven: você informa o recurso (--resource) e o script sempre olha os
snapshots com SnapshotType=manual daquele recurso — com foco nos que sobram de
update/modificação (ex.: snapshots de pré-upgrade), que se acumulam e geram custo.

Workflow:
1. Credenciais da conta bastion do ambiente (AWS_SECRETS ou AWS_*).
2. Para cada conta-alvo (CSV/lista) assume --assume-role e, por região, lista os
   snapshots manuais do(s) recurso(s).
3. Seleciona candidatos por idade (--older-than-days) e, opcionalmente, por nome
   (--update-only usa um regex embutido de snapshots de upgrade/modificação; ou
   --name-regex custom).
4. Por SEGURANÇA o padrão é DRY-RUN: só remove de fato com --apply.
5. Relatório Excel com abas success (processados: deleted/dry-run/skipped) e
   failed (erros de acesso/execução ou de delete).

Snapshots automáticos do RDS não são alvo (não podem ser apagados pela API). O
foco é manual: pré-upgrade/pré-modificação e demais snapshots manuais antigos.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from os import getenv
from pathlib import Path
from typing import List, Optional, Set

import boto3
from botocore.exceptions import BotoCoreError, ClientError

# Snapshots manuais tipicamente criados por upgrade/modificação do RDS.
UPDATE_NAME_REGEX = re.compile(r"(?i)(pre-?upgrade|upgrade|pre-?modif|modif|before-)")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Remove snapshots manuais de RDS (instância/cluster) de um recurso em "
            "múltiplas contas (assume-role). Dry-run por padrão; --apply para deletar."
        )
    )
    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument(
        "--accounts",
        help="Lista de contas separadas por vírgula (ex: 111111111111,222222222222).",
    )
    target.add_argument(
        "--accounts-file",
        type=Path,
        help="Arquivo com um account id por linha.",
    )
    target.add_argument(
        "--accounts-csv",
        type=Path,
        help="Arquivo CSV com a coluna `account_id` (ou `account`, `accountId`, primeira coluna).",
    )
    parser.add_argument(
        "--assume-role",
        required=True,
        help="Nome da role a assumir em cada conta-alvo (ex: OrgBackupAdmin).",
    )
    parser.add_argument(
        "--resource",
        required=True,
        help="Identificador(es) de origem, separados por vírgula (DB instance ou cluster).",
    )
    parser.add_argument(
        "--resource-type",
        choices=["db-instance", "db-cluster"],
        default="db-instance",
        help="Tipo do recurso de origem dos snapshots (padrão: db-instance).",
    )
    parser.add_argument(
        "--region",
        default="us-east-1",
        help="Uma ou mais regiões separadas por vírgula (padrão: us-east-1).",
    )
    parser.add_argument(
        "--older-than-days",
        type=int,
        default=7,
        help="Só considera snapshots com mais de N dias (padrão: 7).",
    )
    parser.add_argument(
        "--update-only",
        action="store_true",
        help="Restringe aos snapshots cujo nome casa o padrão de upgrade/modificação.",
    )
    parser.add_argument(
        "--name-regex",
        help="Regex custom no identificador do snapshot (sobrepõe --update-only).",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Executa a remoção de fato. SEM esta flag, roda em dry-run.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Força dry-run (padrão). Vence --apply se ambos forem passados.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=8,
        help="Quantidade máxima de contas processadas em paralelo.",
    )
    parser.add_argument(
        "--report",
        dest="report",
        help=(
            "Arquivo Excel de saída com abas success/failed "
            "(padrão: rds-manual-snapshot-cleanup-<timestamp>.xlsx)."
        ),
    )
    parser.add_argument(
        "--external-id",
        help="ExternalId opcional para o AssumeRole.",
    )
    parser.add_argument(
        "--role-session-name",
        default="rds-manual-snapshot-cleanup",
        help="SessionName do AssumeRole (padrão: rds-manual-snapshot-cleanup).",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Imprime também a saída detalhada em JSON.",
    )
    return parser.parse_args()


# --------------------------------------------------------------------------- #
# Account loading                                                             #
# --------------------------------------------------------------------------- #
def _account_record(account_id: str) -> dict:
    return {"account": account_id}


def _normalize_header(header: str) -> str:
    return "".join(char for char in header.strip().lower() if char.isalnum())


def _account_id_headers() -> Set[str]:
    return {"account", "accountid", "accountnumber", "id", "conta", "contaid", "numerodaconta"}


def _load_accounts_from_csv(accounts_csv: Path) -> List[dict]:
    with accounts_csv.open(encoding="utf-8-sig", newline="") as handler:
        rows = list(csv.reader(handler))
    if not rows or not rows[0]:
        return []

    headers = [_normalize_header(cell) for cell in rows[0]]
    if any(header in _account_id_headers() for header in headers):
        account_index = next(
            index for index, header in enumerate(headers) if header in _account_id_headers()
        )
        return [
            _account_record(row[account_index].strip())
            for row in rows[1:]
            if len(row) > account_index and row[account_index].strip()
        ]

    return [_account_record(row[0].strip()) for row in rows if row and row[0].strip()]


def _dedupe_accounts(values: List[dict]) -> List[dict]:
    seen: Set[str] = set()
    deduped: List[dict] = []
    for value in values:
        account_id = value["account"]
        if account_id and account_id not in seen:
            seen.add(account_id)
            deduped.append(value)
    return deduped


def _account_load_result(values: List[dict]) -> dict:
    valid_accounts = [value for value in values if value["account"]]
    unique_accounts = _dedupe_accounts(valid_accounts)
    return {
        "accounts": unique_accounts,
        "raw_count": len(values),
        "valid_count": len(valid_accounts),
        "empty_count": len(values) - len(valid_accounts),
        "duplicate_count": len(valid_accounts) - len(unique_accounts),
    }


def load_accounts(
    accounts_csv: Optional[str],
    accounts_file: Optional[Path],
    accounts_csv_file: Optional[Path],
) -> dict:
    if accounts_csv:
        values = [_account_record(acc.strip()) for acc in accounts_csv.split(",")]
    elif accounts_file:
        values = [_account_record(line.strip()) for line in accounts_file.read_text().splitlines()]
    elif accounts_csv_file:
        values = _load_accounts_from_csv(accounts_csv_file)
    else:
        raise ValueError("Informe --accounts, --accounts-file ou --accounts-csv.")

    result = _account_load_result(values)
    if not result["accounts"]:
        raise ValueError("Nenhuma conta válida encontrada.")
    return result


def parse_csv_list(raw: str) -> List[str]:
    items = [item.strip() for item in raw.split(",") if item.strip()]
    deduped: List[str] = []
    for item in items:
        if item not in deduped:
            deduped.append(item)
    return deduped


# --------------------------------------------------------------------------- #
# Source (bastion) credentials                                                #
# --------------------------------------------------------------------------- #
def _read_json_path_or_text(raw: str) -> dict:
    raw = raw.strip()
    if not raw:
        return {}

    path = Path(raw)
    if path.exists():
        raw = path.read_text(encoding="utf-8")

    data = json.loads(raw)
    if not isinstance(data, dict):
        raise ValueError("AWS_SECRETS deve ser um JSON válido de mapa.")
    return data


def _resolve_source_credentials() -> dict:
    aws_secrets = getenv("AWS_SECRETS")
    if aws_secrets:
        secrets = _read_json_path_or_text(aws_secrets)
        return {
            "aws_access_key_id": (
                secrets.get("aws_access_key_id")
                or secrets.get("AccessKeyId")
                or secrets.get("accessKeyId")
                or getenv("AWS_ACCESS_KEY_ID")
            ),
            "aws_secret_access_key": (
                secrets.get("aws_secret_access_key")
                or secrets.get("SecretAccessKey")
                or secrets.get("secretAccessKey")
                or getenv("AWS_SECRET_ACCESS_KEY")
            ),
            "aws_session_token": (
                secrets.get("aws_session_token")
                or secrets.get("SessionToken")
                or secrets.get("sessionToken")
                or getenv("AWS_SESSION_TOKEN")
            ),
            "region": (
                secrets.get("region")
                or secrets.get("aws_region")
                or secrets.get("AWS_REGION")
                or getenv("AWS_REGION")
            ),
        }

    return {
        "aws_access_key_id": getenv("AWS_ACCESS_KEY_ID"),
        "aws_secret_access_key": getenv("AWS_SECRET_ACCESS_KEY"),
        "aws_session_token": getenv("AWS_SESSION_TOKEN"),
        "region": getenv("AWS_REGION") or getenv("AWS_DEFAULT_REGION"),
    }


def build_source_session(region_name: str) -> boto3.Session:
    credentials = _resolve_source_credentials()
    access_key_id = credentials.get("aws_access_key_id")
    secret_access_key = credentials.get("aws_secret_access_key")

    if not access_key_id or not secret_access_key:
        raise ValueError(
            "Credenciais da conta bastion não encontradas. Defina AWS_SECRETS ou "
            "AWS_ACCESS_KEY_ID + AWS_SECRET_ACCESS_KEY no ambiente."
        )

    return boto3.Session(
        aws_access_key_id=access_key_id,
        aws_secret_access_key=secret_access_key,
        aws_session_token=credentials.get("aws_session_token"),
        region_name=credentials.get("region") or region_name,
    )


# --------------------------------------------------------------------------- #
# AWS work per account                                                        #
# --------------------------------------------------------------------------- #
def assume_role_for_account(
    source_session: boto3.Session,
    account_id: str,
    role_name: str,
    role_session_name: str,
    external_id: Optional[str],
    region: str,
) -> boto3.Session:
    sts = source_session.client("sts", region_name=region)
    params = {
        "RoleArn": f"arn:aws:iam::{account_id}:role/{role_name}",
        "RoleSessionName": role_session_name,
    }
    if external_id:
        params["ExternalId"] = external_id

    creds = sts.assume_role(**params)["Credentials"]
    return boto3.Session(
        aws_access_key_id=creds["AccessKeyId"],
        aws_secret_access_key=creds["SecretAccessKey"],
        aws_session_token=creds["SessionToken"],
        region_name=region,
    )


def _list_manual_snapshots(client, resource_type: str, resource: str) -> List[dict]:
    snapshots: List[dict] = []
    if resource_type == "db-cluster":
        paginator = client.get_paginator("describe_db_cluster_snapshots")
        for page in paginator.paginate(DBClusterIdentifier=resource, SnapshotType="manual"):
            for snap in page.get("DBClusterSnapshots", []):
                snapshots.append(
                    {
                        "id": snap["DBClusterSnapshotIdentifier"],
                        "type": snap.get("SnapshotType"),
                        "status": snap.get("Status"),
                        "created": snap.get("SnapshotCreateTime"),
                        "source": snap.get("DBClusterIdentifier"),
                    }
                )
    else:
        paginator = client.get_paginator("describe_db_snapshots")
        for page in paginator.paginate(DBInstanceIdentifier=resource, SnapshotType="manual"):
            for snap in page.get("DBSnapshots", []):
                snapshots.append(
                    {
                        "id": snap["DBSnapshotIdentifier"],
                        "type": snap.get("SnapshotType"),
                        "status": snap.get("Status"),
                        "created": snap.get("SnapshotCreateTime"),
                        "source": snap.get("DBInstanceIdentifier"),
                    }
                )
    return snapshots


def _delete_snapshot(client, resource_type: str, snapshot_id: str) -> None:
    if resource_type == "db-cluster":
        client.delete_db_cluster_snapshot(DBClusterSnapshotIdentifier=snapshot_id)
    else:
        client.delete_db_snapshot(DBSnapshotIdentifier=snapshot_id)


def _age_days(created) -> Optional[int]:
    if not created:
        return None
    now = datetime.now(timezone.utc)
    return (now - created).days


def _name_filter(args: argparse.Namespace) -> Optional[re.Pattern]:
    if args.name_regex:
        return re.compile(args.name_regex)
    if args.update_only:
        return UPDATE_NAME_REGEX
    return None


def _evaluate_snapshot(snap: dict, args: argparse.Namespace, name_re: Optional[re.Pattern]) -> dict:
    age = _age_days(snap["created"])
    row = {
        "region": None,
        "source": snap["source"],
        "snapshot_id": snap["id"],
        "snapshot_type": snap["type"],
        "created": snap["created"].isoformat() if snap["created"] else "",
        "age_days": age if age is not None else "",
        "action": "skipped",
        "reason": "",
        "error": None,
    }
    if snap["status"] != "available":
        row["reason"] = f"status:{snap['status']}"
    elif age is None or age < args.older_than_days:
        row["reason"] = "too-recent"
    elif name_re and not name_re.search(snap["id"]):
        row["reason"] = "name-no-match"
    else:
        row["action"] = "candidate"
    return row


def _extract_error_code(error: Exception) -> str:
    if not isinstance(error, ClientError):
        return ""
    return (error.response or {}).get("Error", {}).get("Code", "")


def log_step(message: str) -> None:
    print(f"[{datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}] {message}", file=sys.stderr, flush=True)


def check_account(
    index: int,
    account: dict,
    args: argparse.Namespace,
    source_session: boto3.Session,
    regions: List[str],
    resources: List[str],
    dry_run: bool,
) -> tuple[int, dict]:
    account_id = account["account"]
    name_re = _name_filter(args)
    result = {
        "account": account_id,
        "ok": False,
        "error": None,
        "error_code": None,
        "snapshots": [],
    }
    try:
        log_step(f"Conta {index + 1}: iniciando {account_id}")
        assumed_session = assume_role_for_account(
            source_session=source_session,
            account_id=account_id,
            role_name=args.assume_role,
            role_session_name=args.role_session_name,
            external_id=args.external_id,
            region=regions[0],
        )
        log_step(f"Conta {account_id}: assumeRole concluido")

        rows: List[dict] = []
        for region in regions:
            rds = assumed_session.client("rds", region_name=region)
            for resource in resources:
                for snap in _list_manual_snapshots(rds, args.resource_type, resource):
                    row = _evaluate_snapshot(snap, args, name_re)
                    row["region"] = region
                    row["account"] = account_id
                    if row["action"] == "candidate":
                        if dry_run:
                            row["action"] = "dry-run"
                        else:
                            try:
                                _delete_snapshot(rds, args.resource_type, snap["id"])
                                row["action"] = "deleted"
                            except (ClientError, BotoCoreError) as error:
                                row["action"] = "failed"
                                row["error"] = str(error)
                    rows.append(row)
            log_step(
                f"Conta {account_id} [{region}]: snapshots={sum(1 for r in rows if r['region'] == region)}"
            )
        result["snapshots"] = rows
        result["ok"] = True
    except (ClientError, BotoCoreError, ValueError) as error:
        result["error_code"] = _extract_error_code(error) or None
        result["error"] = str(error)
        result["ok"] = False
        log_step(f"Conta {account_id}: erro: {result['error']}")
    return index, result


# --------------------------------------------------------------------------- #
# Reporting                                                                   #
# --------------------------------------------------------------------------- #
def success_headers() -> List[str]:
    return ["account", "region", "source", "snapshot_id", "snapshot_type", "created", "age_days", "action", "reason"]


def success_row(item: dict) -> List[str]:
    return [
        item["account"],
        item["region"],
        item["source"] or "",
        item["snapshot_id"],
        item["snapshot_type"] or "",
        item["created"],
        str(item["age_days"]),
        item["action"],
        item["reason"],
    ]


def failed_headers() -> List[str]:
    return ["account", "region", "snapshot_id", "error_code", "error"]


def all_snapshots(results: List[dict]) -> List[dict]:
    rows: List[dict] = []
    for item in results:
        if item["ok"]:
            rows.extend(item["snapshots"])
    return rows


def write_xlsx_report(report_path: str, results: List[dict]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as error:
        raise ValueError("Para gerar relatório Excel, instale openpyxl.") from error

    snapshots = all_snapshots(results)
    success = [row for row in snapshots if row["error"] is None]
    snapshot_failures = [row for row in snapshots if row["error"] is not None]
    account_failures = [item for item in results if not item["ok"]]

    workbook = Workbook()
    workbook.remove(workbook.active)

    success_sheet = workbook.create_sheet("success")
    success_sheet.append(success_headers())
    for row in success:
        success_sheet.append(success_row(row))

    failed_sheet = workbook.create_sheet("failed")
    failed_sheet.append(failed_headers())
    for item in account_failures:
        failed_sheet.append([item["account"], "", "", item["error_code"] or "", item["error"] or ""])
    for row in snapshot_failures:
        failed_sheet.append([row["account"], row["region"], row["snapshot_id"], "", row["error"] or ""])

    workbook.save(report_path)


def normalize_report_path(report_path: str) -> str:
    path = Path(report_path)
    if path.suffix.lower() == ".xlsx":
        return str(path)
    return str(path.with_suffix(".xlsx"))


def write_reports(report_path: str, results: List[dict]) -> str:
    report_path = normalize_report_path(report_path)
    write_xlsx_report(report_path, results)

    snapshots = all_snapshots(results)
    deleted = sum(1 for row in snapshots if row["action"] == "deleted")
    planned = sum(1 for row in snapshots if row["action"] == "dry-run")
    skipped = sum(1 for row in snapshots if row["action"] == "skipped")
    snap_failed = sum(1 for row in snapshots if row["action"] == "failed")
    acct_failed = sum(1 for item in results if not item["ok"])
    log_step(
        f"Relatorio Excel gerado: {report_path} sheets=success,failed "
        f"deleted={deleted} dry_run={planned} skipped={skipped} "
        f"snapshot_falhas={snap_failed} conta_falhas={acct_failed}"
    )
    return report_path


# --------------------------------------------------------------------------- #
# Orchestration                                                               #
# --------------------------------------------------------------------------- #
def run_check(args: argparse.Namespace) -> int:
    if args.workers <= 0:
        raise ValueError("workers precisa ser maior que 0.")
    if args.older_than_days < 0:
        raise ValueError("older-than-days não pode ser negativo.")

    # Dry-run é o padrão seguro; --apply executa; --dry-run vence se ambos vierem.
    dry_run = True if args.dry_run else not args.apply

    regions = parse_csv_list(args.region) or ["us-east-1"]
    resources = parse_csv_list(args.resource)
    if not resources:
        raise ValueError("Informe ao menos um recurso em --resource.")

    account_load = load_accounts(args.accounts, args.accounts_file, args.accounts_csv)
    accounts = account_load["accounts"]
    source_session = build_source_session(regions[0])
    workers = min(args.workers, len(accounts))

    log_step(
        "Carga de contas: "
        f"entradas_lidas={account_load['raw_count']} "
        f"validas={account_load['valid_count']} "
        f"vazias_ignoradas={account_load['empty_count']} "
        f"duplicadas_removidas={account_load['duplicate_count']} "
        f"unicas_processadas={len(accounts)}"
    )
    modo = "DRY-RUN (nada será apagado)" if dry_run else "APPLY (vai apagar de fato)"
    log_step(
        f"Limpeza de snapshots manuais [{modo}]: contas={len(accounts)} workers={workers} "
        f"assume_role={args.assume_role} tipo={args.resource_type} "
        f"recursos={','.join(resources)} regioes={','.join(regions)} "
        f"older_than_days={args.older_than_days} "
        f"filtro_nome={'update-only' if args.update_only and not args.name_regex else (args.name_regex or 'nenhum')}"
    )

    results: list[dict] = [None for _ in accounts]  # type: ignore[list-item]
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [
            executor.submit(
                check_account,
                index=index,
                account=account,
                args=args,
                source_session=source_session,
                regions=regions,
                resources=resources,
                dry_run=dry_run,
            )
            for index, account in enumerate(accounts)
        ]
        for future in as_completed(futures):
            index, result = future.result()
            results[index] = result

    report_path = args.report or f"rds-manual-snapshot-cleanup-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}.xlsx"
    write_reports(report_path, results)

    snapshots = all_snapshots(results)
    deleted = [row for row in snapshots if row["action"] == "deleted"]
    planned = [row for row in snapshots if row["action"] == "dry-run"]
    snap_failed = [row for row in snapshots if row["action"] == "failed"]
    acct_failed = [item for item in results if not item["ok"]]

    if args.json:
        print(json.dumps(results, indent=2, ensure_ascii=False, default=str))
    else:
        for item in results:
            if not item["ok"]:
                print(f'{item["account"]}: ERRO - {item["error"] or "erro desconhecido"}')
                continue
            d = sum(1 for r in item["snapshots"] if r["action"] == "deleted")
            p = sum(1 for r in item["snapshots"] if r["action"] == "dry-run")
            s = sum(1 for r in item["snapshots"] if r["action"] == "skipped")
            f = sum(1 for r in item["snapshots"] if r["action"] == "failed")
            print(f'{item["account"]}: OK - deleted={d} dry_run={p} skipped={s} failed={f}')

    log_step(
        f"Concluido: deleted={len(deleted)} dry_run={len(planned)} "
        f"snapshot_falhas={len(snap_failed)} conta_falhas={len(acct_failed)}"
    )

    if acct_failed or snap_failed:
        return 2
    return 0


def main() -> int:
    args = parse_args()
    return run_check(args)


if __name__ == "__main__":
    raise SystemExit(main())
